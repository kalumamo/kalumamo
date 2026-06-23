#!/usr/bin/env python
"""
Create test dataset with 3 products showing clear performance differences.
Products: HIGH, MEDIUM, LOW performers with different scores and alerts.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Create three distinct test products with clear performance differences
products = [
    {
        "name": "MOBILE_HIGH",
        "code": "MOBILE_01",
        "scenario": "HIGH PERFORMER",
        "total_users": 850000,
        "active_users": 680000,  # 80% engagement
        "new_users": 12000,
        "churned_users": 2000,
        "total_transactions": 3500000,
        "successful_transactions": 3360000,  # 96% success
        "failed_transactions": 140000,
        "transaction_volume": 8500000000,
        "total_revenue": 85000000,
        "fee_revenue": 5100000,
        "uptime_percentage": 99.8,  # Excellent uptime
        "downtime_hours": 5,
        "downtime_minutes": 288,
        "avg_response_time_ms": 250,  # Fast
        "api_error_rate": 0.5,  # Low errors
        "total_complaints": 45,
        "resolved_complaints": 42,
        "csat_score": 4.7,  # High satisfaction
        "fraud_event_count": 2,
        "security_incident_count": 0,
    },
    {
        "name": "CARD_MEDIUM",
        "code": "CARD_01",
        "scenario": "MEDIUM PERFORMER",
        "total_users": 550000,
        "active_users": 319000,  # 58% engagement
        "new_users": 6500,
        "churned_users": 5500,
        "total_transactions": 1800000,
        "successful_transactions": 1620000,  # 90% success
        "failed_transactions": 180000,
        "transaction_volume": 5000000000,
        "total_revenue": 50000000,
        "fee_revenue": 2500000,
        "uptime_percentage": 97.5,  # Good uptime
        "downtime_hours": 18,
        "downtime_minutes": 1080,
        "avg_response_time_ms": 420,  # Moderate
        "api_error_rate": 2.2,  # Moderate errors
        "total_complaints": 120,
        "resolved_complaints": 96,  # 80% resolution
        "csat_score": 3.8,  # Decent satisfaction
        "fraud_event_count": 8,
        "security_incident_count": 1,
    },
    {
        "name": "ATM_LOW",
        "code": "ATM_01",
        "scenario": "LOW PERFORMER - CRISIS",
        "total_users": 320000,
        "active_users": 128000,  # 40% engagement - poor
        "new_users": 1600,
        "churned_users": 12800,  # High churn
        "total_transactions": 850000,
        "successful_transactions": 680000,  # 80% success - low
        "failed_transactions": 170000,
        "transaction_volume": 2100000000,
        "total_revenue": 21000000,
        "fee_revenue": 1050000,
        "uptime_percentage": 94.2,  # Poor uptime
        "downtime_hours": 136,  # High downtime
        "downtime_minutes": 8160,
        "avg_response_time_ms": 850,  # Slow
        "api_error_rate": 6.8,  # High errors
        "total_complaints": 385,
        "resolved_complaints": 270,  # 70% resolution - low
        "csat_score": 2.9,  # Low satisfaction
        "fraud_event_count": 24,
        "security_incident_count": 3,
    }
]

# Create dataframe
data = []
period_date = datetime.now().date()

for product in products:
    row = {
        "product_code": product["code"],
        "period_date": str(period_date),
        "total_users": product["total_users"],
        "active_users": product["active_users"],
        "new_users": product["new_users"],
        "churned_users": product["churned_users"],
        "total_transactions": product["total_transactions"],
        "successful_transactions": product["successful_transactions"],
        "failed_transactions": product["failed_transactions"],
        "transaction_volume": product["transaction_volume"],
        "total_revenue": product["total_revenue"],
        "fee_revenue": product["fee_revenue"],
        "uptime_percentage": product["uptime_percentage"],
        "downtime_hours": product["downtime_hours"],
        "downtime_minutes": product["downtime_minutes"],
        "avg_response_time_ms": product["avg_response_time_ms"],
        "api_error_rate": product["api_error_rate"],
        "total_complaints": product["total_complaints"],
        "resolved_complaints": product["resolved_complaints"],
        "csat_score": product["csat_score"],
        "fraud_event_count": product["fraud_event_count"],
        "security_incident_count": product["security_incident_count"],
    }
    data.append(row)

df = pd.DataFrame(data)

# Create Excel workbook with multiple sheets
wb = Workbook()
ws_upload = wb.active
ws_upload.title = "Upload_Ready"

# Add title and description
ws_upload['A1'] = "AHADU PULSE — TEST DATASET (3 Products: High/Medium/Low)"
ws_upload['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_upload['A1'].fill = PatternFill(start_color="9B1535", end_color="9B1535", fill_type="solid")
ws_upload.merge_cells('A1:V1')

ws_upload['A2'] = "Upload this sheet to see clear scores, rankings, alerts, and recommendations"
ws_upload['A2'].font = Font(italic=True, size=10, color="666666")
ws_upload.merge_cells('A2:V2')

# Headers (row 3)
headers = list(df.columns)
for col_idx, header in enumerate(headers, 1):
    cell = ws_upload.cell(row=3, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="7A0E28", end_color="7A0E28", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Add data with colors for each product
colors = {
    "MOBILE_01": "C6EFCE",  # Green for HIGH
    "CARD_01": "FFEB9C",    # Yellow for MEDIUM
    "ATM_01": "FFC7CE",     # Red for LOW
}

for row_idx, (_, row_data) in enumerate(df.iterrows(), 4):
    product_code = row_data["product_code"]
    color = colors.get(product_code, "FFFFFF")
    
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_upload.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(value, (int, float)):
            if isinstance(value, float):
                cell.number_format = '0.00'
            else:
                cell.number_format = '0'

# Add legend/notes sheet
ws_notes = wb.create_sheet("Legend")
ws_notes['A1'] = "TEST DATASET SCENARIOS"
ws_notes['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_notes['A1'].fill = PatternFill(start_color="9B1535", end_color="9B1535", fill_type="solid")
ws_notes.merge_cells('A1:D1')

notes_data = [
    ["Product", "Scenario", "Expected Score", "Expected Tier", "Key Metrics"],
    [
        "MOBILE_01",
        "HIGH PERFORMER",
        "85-95",
        "HIGH ✓",
        "96% txn success, 99.8% uptime, 4.7 CSAT, 80% engagement"
    ],
    [
        "CARD_01",
        "MEDIUM PERFORMER",
        "65-75",
        "MEDIUM",
        "90% txn success, 97.5% uptime, 3.8 CSAT, 58% engagement"
    ],
    [
        "ATM_01",
        "LOW PERFORMER - CRISIS",
        "35-45",
        "LOW ⚠",
        "80% txn success, 94.2% uptime, 2.9 CSAT, 40% engagement, HIGH DOWNTIME"
    ]
]

for row_idx, note_row in enumerate(notes_data, 3):
    for col_idx, value in enumerate(note_row, 1):
        cell = ws_notes.cell(row=row_idx, column=col_idx)
        cell.value = value
        if row_idx == 3:  # Headers
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7A0E28", end_color="7A0E28", fill_type="solid")
        else:
            if "HIGH" in value:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif "MEDIUM" in value:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif "LOW" in value:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Add expected outputs sheet
ws_expected = wb.create_sheet("Expected_Outputs")
ws_expected['A1'] = "EXPECTED SYSTEM OUTPUTS"
ws_expected['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_expected['A1'].fill = PatternFill(start_color="9B1535", end_color="9B1535", fill_type="solid")
ws_expected.merge_cells('A1:E1')

expected_outputs = [
    ["Product", "Score", "Tier", "Alert Severity", "Sample Recommendation"],
    [
        "MOBILE_01",
        "88-92",
        "HIGH",
        "None (Optimal)",
        "Maintain current operational standards"
    ],
    [
        "CARD_01",
        "70",
        "MEDIUM",
        "Warning: Uptime declining",
        "Investigate API performance issues, improve infrastructure"
    ],
    [
        "ATM_01",
        "42",
        "LOW",
        "CRITICAL: Service degradation",
        "URGENT: Resolve ATM hardware issues, implement maintenance plan"
    ]
]

for row_idx, output_row in enumerate(expected_outputs, 3):
    for col_idx, value in enumerate(output_row, 1):
        cell = ws_expected.cell(row=row_idx, column=col_idx)
        cell.value = value
        if row_idx == 3:  # Headers
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7A0E28", end_color="7A0E28", fill_type="solid")
        else:
            if "HIGH" in value:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif "MEDIUM" in value:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif "LOW" in value or "CRITICAL" in value:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Adjust column widths
for ws in [ws_upload, ws_notes, ws_expected]:
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

# Save file
output_path = "TEST_DATASET_3_PRODUCTS.xlsx"
wb.save(output_path)
print(f"✓ Created: {output_path}")
print(f"\nDataset contains 3 products for testing:")
print(f"  1. MOBILE_01 - HIGH PERFORMER (Score ~90, Tier: HIGH) ✓")
print(f"  2. CARD_01  - MEDIUM PERFORMER (Score ~70, Tier: MEDIUM)")
print(f"  3. ATM_01   - LOW PERFORMER (Score ~42, Tier: LOW) ⚠ CRITICAL ALERTS")
print(f"\nUpload Instructions:")
print(f"  1. Open Settings → Upload Data")
print(f"  2. Select {output_path}")
print(f"  3. Watch the automatic processing:")
print(f"     - Features computed")
print(f"     - Scores calculated")
print(f"     - Alerts generated")
print(f"     - Recommendations created")
print(f"  4. Go to Dashboard to see:")
print(f"     - Rankings page → 3 products ranked by score")
print(f"     - Scores page → Individual product scores")
print(f"     - Alerts page → All alerts (especially for ATM_01)")
print(f"     - Recommendations page → AI-generated actions")
