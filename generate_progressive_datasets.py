#!/usr/bin/env python
"""
Generate 4 progressive datasets showing data changes over time.
Each dataset shows products improving, degrading, or changing.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

# Dataset 1: Initial baseline (June 22, 2026)
dataset1 = {
    "name": "DATASET_1_INITIAL.xlsx",
    "date": "2026-06-22",
    "scenario": "Initial Baseline",
    "products": [
        {
            "code": "MOBILE_01",
            "total_users": 850000,
            "active_users": 680000,
            "new_users": 12000,
            "churned_users": 2000,
            "total_transactions": 3500000,
            "successful_transactions": 3360000,
            "failed_transactions": 140000,
            "transaction_volume": 8500000000,
            "total_revenue": 85000000,
            "fee_revenue": 5100000,
            "uptime_percentage": 99.8,
            "downtime_hours": 5,
            "downtime_minutes": 288,
            "avg_response_time_ms": 250,
            "api_error_rate": 0.5,
            "total_complaints": 45,
            "resolved_complaints": 42,
            "csat_score": 4.7,
            "fraud_event_count": 2,
            "security_incident_count": 0,
        },
        {
            "code": "CARD_01",
            "total_users": 550000,
            "active_users": 319000,
            "new_users": 6500,
            "churned_users": 5500,
            "total_transactions": 1800000,
            "successful_transactions": 1620000,
            "failed_transactions": 180000,
            "transaction_volume": 5000000000,
            "total_revenue": 50000000,
            "fee_revenue": 2500000,
            "uptime_percentage": 97.5,
            "downtime_hours": 18,
            "downtime_minutes": 1080,
            "avg_response_time_ms": 420,
            "api_error_rate": 2.2,
            "total_complaints": 120,
            "resolved_complaints": 96,
            "csat_score": 3.8,
            "fraud_event_count": 8,
            "security_incident_count": 1,
        },
        {
            "code": "ATM_01",
            "total_users": 320000,
            "active_users": 128000,
            "new_users": 1600,
            "churned_users": 12800,
            "total_transactions": 850000,
            "successful_transactions": 680000,
            "failed_transactions": 170000,
            "transaction_volume": 2100000000,
            "total_revenue": 21000000,
            "fee_revenue": 1050000,
            "uptime_percentage": 94.2,
            "downtime_hours": 136,
            "downtime_minutes": 8160,
            "avg_response_time_ms": 850,
            "api_error_rate": 6.8,
            "total_complaints": 385,
            "resolved_complaints": 270,
            "csat_score": 2.9,
            "fraud_event_count": 24,
            "security_incident_count": 3,
        },
    ]
}

# Dataset 2: One week later - MOBILE improves, CARD stays, ATM worsens
dataset2 = {
    "name": "DATASET_2_WEEK2.xlsx",
    "date": "2026-06-29",
    "scenario": "Week 2 - Changes Observed",
    "products": [
        {
            "code": "MOBILE_01",
            "total_users": 862000,      # +12K
            "active_users": 698000,     # +18K (better engagement)
            "new_users": 14000,         # +2K
            "churned_users": 1200,      # -800 (less churn)
            "total_transactions": 3650000,  # +150K
            "successful_transactions": 3555500,  # +195.5K (better)
            "failed_transactions": 94500,  # -45.5K (fewer failures)
            "transaction_volume": 8900000000,  # +400M
            "total_revenue": 89000000,  # +4M (revenue growth)
            "fee_revenue": 5340000,     # +240K
            "uptime_percentage": 99.85, # +0.05 (even better)
            "downtime_hours": 3,        # -2
            "downtime_minutes": 180,    # -108
            "avg_response_time_ms": 235,  # -15 (faster)
            "api_error_rate": 0.3,      # -0.2 (fewer errors)
            "total_complaints": 32,     # -13
            "resolved_complaints": 31,  # -11
            "csat_score": 4.8,          # +0.1 (higher)
            "fraud_event_count": 1,     # -1 (less fraud)
            "security_incident_count": 0,  # no change
        },
        {
            "code": "CARD_01",
            "total_users": 551000,      # +1K (stable)
            "active_users": 321000,     # +2K (slight improvement)
            "new_users": 6600,          # +100
            "churned_users": 5400,      # -100
            "total_transactions": 1810000,  # +10K
            "successful_transactions": 1630000,  # +10K (stable)
            "failed_transactions": 180000,  # no change
            "transaction_volume": 5050000000,  # +50M
            "total_revenue": 50500000,  # +500K
            "fee_revenue": 2525000,     # +25K
            "uptime_percentage": 97.6,  # +0.1 (slight improvement)
            "downtime_hours": 17,       # -1
            "downtime_minutes": 1020,   # -60
            "avg_response_time_ms": 418,  # -2 (slightly faster)
            "api_error_rate": 2.1,      # -0.1 (improving)
            "total_complaints": 118,    # -2
            "resolved_complaints": 97,  # +1
            "csat_score": 3.85,         # +0.05 (slight improvement)
            "fraud_event_count": 7,     # -1
            "security_incident_count": 1,  # no change
        },
        {
            "code": "ATM_01",
            "total_users": 315000,      # -5K (losing users)
            "active_users": 118000,     # -10K (engagement drop)
            "new_users": 1200,          # -400
            "churned_users": 14000,     # +1.2K (more churn)
            "total_transactions": 800000,  # -50K (fewer transactions)
            "successful_transactions": 620000,  # -60K (worse)
            "failed_transactions": 180000,  # +10K (more failures)
            "transaction_volume": 1950000000,  # -150M
            "total_revenue": 19500000,  # -1.5M (revenue loss)
            "fee_revenue": 975000,      # -75K
            "uptime_percentage": 93.5,  # -0.7 (worse)
            "downtime_hours": 155,      # +19 (more downtime)
            "downtime_minutes": 9300,   # +1140
            "avg_response_time_ms": 920,  # +70 (slower)
            "api_error_rate": 7.5,      # +0.7 (worse)
            "total_complaints": 420,    # +35 (more complaints)
            "resolved_complaints": 280,  # +10
            "csat_score": 2.65,         # -0.25 (worse)
            "fraud_event_count": 32,    # +8 (more fraud)
            "security_incident_count": 4,  # +1
        },
    ]
}

# Dataset 3: Two weeks later - More divergence
dataset3 = {
    "name": "DATASET_3_WEEK3.xlsx",
    "date": "2026-07-06",
    "scenario": "Week 3 - Trends Accelerating",
    "products": [
        {
            "code": "MOBILE_01",
            "total_users": 878000,      # +16K more
            "active_users": 720000,     # +22K more
            "new_users": 16500,         # +2.5K
            "churned_users": 900,       # -300 more
            "total_transactions": 3850000,  # +200K
            "successful_transactions": 3771500,  # +216K (96.9% success!)
            "failed_transactions": 78500,  # -16K
            "transaction_volume": 9400000000,  # +500M more
            "total_revenue": 94000000,  # +5M more (strong growth)
            "fee_revenue": 5640000,     # +300K
            "uptime_percentage": 99.9,  # +0.05 (near perfect!)
            "downtime_hours": 2,        # -1 more
            "downtime_minutes": 120,    # -60
            "avg_response_time_ms": 220,  # -15 more
            "api_error_rate": 0.2,      # -0.1 more
            "total_complaints": 25,     # -7 more
            "resolved_complaints": 24,  # -7 more
            "csat_score": 4.85,         # +0.05 (continuing improvement)
            "fraud_event_count": 0,     # -1 (zero fraud!)
            "security_incident_count": 0,
        },
        {
            "code": "CARD_01",
            "total_users": 548000,      # -3K (starting to decline)
            "active_users": 315000,     # -6K (engagement dropping)
            "new_users": 6200,          # -400
            "churned_users": 5800,      # +400
            "total_transactions": 1780000,  # -30K (declining)
            "successful_transactions": 1602000,  # -28K (worse)
            "failed_transactions": 178000,  # -2K
            "transaction_volume": 4900000000,  # -150M (decline)
            "total_revenue": 49000000,  # -1.5M (revenue drop)
            "fee_revenue": 2450000,     # -75K
            "uptime_percentage": 97.2,  # -0.4 (declining)
            "downtime_hours": 20,       # +3
            "downtime_minutes": 1200,   # +180
            "avg_response_time_ms": 440,  # +22 (slower)
            "api_error_rate": 2.5,      # +0.4 (worse)
            "total_complaints": 135,    # +17
            "resolved_complaints": 100,  # +3
            "csat_score": 3.7,          # -0.15 (declining)
            "fraud_event_count": 10,    # +3
            "security_incident_count": 2,  # +1
        },
        {
            "code": "ATM_01",
            "total_users": 305000,      # -10K more
            "active_users": 100000,     # -18K more (crisis)
            "new_users": 800,           # -400 more
            "churned_users": 16000,     # +2K (mass churn)
            "total_transactions": 720000,  # -80K (crisis)
            "successful_transactions": 518400,  # -101.6K (72% success - crisis)
            "failed_transactions": 201600,  # +21.6K
            "transaction_volume": 1750000000,  # -200M (decline)
            "total_revenue": 17500000,  # -2M (significant loss)
            "fee_revenue": 875000,      # -100K
            "uptime_percentage": 92.0,  # -1.5 (CRITICAL!)
            "downtime_hours": 192,      # +37 (much worse)
            "downtime_minutes": 11520,  # +2220
            "avg_response_time_ms": 1050,  # +130 (much slower)
            "api_error_rate": 9.2,      # +1.7 (much worse)
            "total_complaints": 510,    # +90 (crisis)
            "resolved_complaints": 310,  # +30
            "csat_score": 2.2,          # -0.45 (crisis level)
            "fraud_event_count": 48,    # +16
            "security_incident_count": 6,  # +2
        },
    ]
}

# Dataset 4: Three weeks later - Final status
dataset4 = {
    "name": "DATASET_4_WEEK4.xlsx",
    "date": "2026-07-13",
    "scenario": "Week 4 - Stabilizing",
    "products": [
        {
            "code": "MOBILE_01",
            "total_users": 895000,      # +17K (continued growth)
            "active_users": 745000,     # +25K (strong engagement)
            "new_users": 18000,         # +1.5K
            "churned_users": 700,       # -200 (low churn)
            "total_transactions": 4050000,  # +200K
            "successful_transactions": 3969000,  # +197.5K (98% success!)
            "failed_transactions": 81000,  # -2.5K
            "transaction_volume": 9900000000,  # +500M
            "total_revenue": 99000000,  # +5M (reaching 100M milestone)
            "fee_revenue": 5940000,     # +300K
            "uptime_percentage": 99.95, # +0.05 (near perfect)
            "downtime_hours": 1,        # -1 (minimal downtime)
            "downtime_minutes": 60,     # -60
            "avg_response_time_ms": 210,  # -10 (fastest)
            "api_error_rate": 0.1,      # -0.1 (minimal errors)
            "total_complaints": 18,     # -7
            "resolved_complaints": 17,  # -7
            "csat_score": 4.9,          # +0.05 (highest possible)
            "fraud_event_count": 0,     # no change
            "security_incident_count": 0,
        },
        {
            "code": "CARD_01",
            "total_users": 540000,      # -8K total (declining)
            "active_users": 302000,     # -13K (significant drop)
            "new_users": 5500,          # -700
            "churned_users": 6500,      # +700
            "total_transactions": 1720000,  # -60K
            "successful_transactions": 1548000,  # -54K (90% success)
            "failed_transactions": 172000,  # -6K
            "transaction_volume": 4700000000,  # -200M
            "total_revenue": 47000000,  # -2M
            "fee_revenue": 2350000,     # -100K
            "uptime_percentage": 96.8,  # -0.4 (declining trend)
            "downtime_hours": 23,       # +3
            "downtime_minutes": 1380,   # +180
            "avg_response_time_ms": 465,  # +25
            "api_error_rate": 2.8,      # +0.3
            "total_complaints": 155,    # +20
            "resolved_complaints": 112,  # +12
            "csat_score": 3.55,         # -0.15 (declining)
            "fraud_event_count": 12,    # +2
            "security_incident_count": 2,
        },
        {
            "code": "ATM_01",
            "total_users": 298000,      # -7K (continued decline)
            "active_users": 93500,      # -6.5K (worsening)
            "new_users": 600,           # -200
            "churned_users": 17500,     # +1.5K (continued churn)
            "total_transactions": 670000,  # -50K (continuing crisis)
            "successful_transactions": 483600,  # -34.8K (72.2% - crisis)
            "failed_transactions": 186400,  # +15.2K
            "transaction_volume": 1600000000,  # -150M
            "total_revenue": 16000000,  # -1.5M
            "fee_revenue": 800000,      # -75K
            "uptime_percentage": 90.5,  # -1.5 (critical deterioration!)
            "downtime_hours": 224,      # +32 (serious)
            "downtime_minutes": 13440,  # +1920
            "avg_response_time_ms": 1200,  # +150 (very slow)
            "api_error_rate": 10.5,     # +1.3 (critical)
            "total_complaints": 580,    # +70
            "resolved_complaints": 340,  # +30
            "csat_score": 1.95,         # -0.25 (critical)
            "fraud_event_count": 56,    # +8
            "security_incident_count": 7,  # +1
        },
    ]
}

datasets = [dataset1, dataset2, dataset3, dataset4]

def create_dataset_file(dataset_info):
    """Create an Excel file for a dataset with proper formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Upload_Ready"
    
    # Title
    ws['A1'] = f"AHADU PULSE — {dataset_info['scenario']} ({dataset_info['date']})"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="9B1535", end_color="9B1535", fill_type="solid")
    ws.merge_cells('A1:V1')
    
    ws['A2'] = "Upload this dataset to see updated scores, rankings, alerts, and recommendations"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:V2')
    
    # Create dataframe
    data = []
    for product in dataset_info['products']:
        row = {
            "product_code": product["code"],
            "period_date": dataset_info['date'],
            "total_users": product["total_users"],
            "active_users": product["active_users"],
            "new_users": product["new_users"],
            "churned_users": product["churned_users"],
            "total_transactions": product["total_transactions"],
            "successful_transactions": product["successful_transactions"],
            "failed_transactions": product["failed_transactions"],
            "transaction_volume": product["transaction_volume"],
            "total_revenue": product["total_revenue"],
            "fee_revenue": product["fee_revenue"],
            "uptime_percentage": product["uptime_percentage"],
            "downtime_hours": product["downtime_hours"],
            "downtime_minutes": product["downtime_minutes"],
            "avg_response_time_ms": product["avg_response_time_ms"],
            "api_error_rate": product["api_error_rate"],
            "total_complaints": product["total_complaints"],
            "resolved_complaints": product["resolved_complaints"],
            "csat_score": product["csat_score"],
            "fraud_event_count": product["fraud_event_count"],
            "security_incident_count": product["security_incident_count"],
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    
    # Headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="7A0E28", end_color="7A0E28", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Color mapping
    colors = {
        "MOBILE_01": "C6EFCE",  # Green (improving)
        "CARD_01": "FFEB9C",    # Yellow (medium)
        "ATM_01": "FFC7CE",     # Red (critical)
    }
    
    # Data rows
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 4):
        product_code = row_data["product_code"]
        color = colors.get(product_code, "FFFFFF")
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(value, (int, float)):
                if isinstance(value, float):
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '0'
    
    # Adjust widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save
    output_path = dataset_info['name']
    wb.save(output_path)
    print(f"✓ Created: {output_path} ({dataset_info['scenario']})")

# Generate all 4 datasets
print("=" * 80)
print("GENERATING 4 PROGRESSIVE DATASETS")
print("=" * 80)
print()

for dataset_info in datasets:
    create_dataset_file(dataset_info)

print()
print("=" * 80)
print("UPLOAD INSTRUCTIONS")
print("=" * 80)
print()
print("Upload datasets in this order to see data changes:")
print()
print("1. DATASET_1_INITIAL.xlsx (Baseline - June 22)")
print("   → MOBILE: Score 90 (HIGH) | CARD: Score 70 (MEDIUM) | ATM: Score 42 (LOW)")
print()
print("2. DATASET_2_WEEK2.xlsx (Week 2 - June 29)")
print("   → MOBILE: Improving (92+) | CARD: Stable (~70) | ATM: Worsening (40-)")
print()
print("3. DATASET_3_WEEK3.xlsx (Week 3 - July 6)")
print("   → MOBILE: Excellent (94+) | CARD: Declining (68-) | ATM: Crisis (35-)")
print()
print("4. DATASET_4_WEEK4.xlsx (Week 4 - July 13)")
print("   → MOBILE: Peak (96+) | CARD: Poor (65-) | ATM: Critical (30)")
print()
print("=" * 80)
print("EXPECTED BEHAVIOR")
print("=" * 80)
print()
print("Each upload will:")
print("  ✓ Replace old data with new data")
print("  ✓ Recalculate scores")
print("  ✓ Update rankings with new scores")
print("  ✓ Generate new alerts based on new metrics")
print("  ✓ Create new recommendations")
print("  ✓ Refresh all dashboard pages")
print()
print("You should see:")
print("  • Dashboard: Average score changes")
print("  • Rankings: Products move up/down based on new scores")
print("  • Scores: Each product's score increases/decreases")
print("  • Alerts: New alerts appear or old ones resolve")
print("  • Recommendations: Different recommendations for each status")
print()
print("=" * 80)
